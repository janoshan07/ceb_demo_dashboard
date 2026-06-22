# Python script to generate a CEB billing test workbook using openpyxl
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Create workbook
wb = openpyxl.Workbook()

# Define headers
headers = [
    "Account No", "Customer Name", "Bank Code", "Branch Code", 
    "Bank Account No", "Ref No", "From Date", "To Date", 
    "Imports", "Exports", "Unit Cost", "Exp Code"
]

# Sheet data configurations
sheets_data = {
    "Colombo West": [
        # Valid billing records
        ["1002345091", "Sun Industrial Pvt Ltd", "BOC", "032", "7045920", "REF-202606-01", "2026-06-01", "2026-06-30", 1100, 6400, 48.0, "EXP-IND-A"],
        ["1005520938", "Siddhalepa Hospital", "HNB", "014", "8829402", "REF-202606-02", "2026-06-01", "2026-06-30", 3200, 4100, 52.0, "EXP-MED-B"],
        ["3001204092", "Keells Supermarket Col 3", "SAMP", "087", "8876529", "REF-202606-03", "2026-06-01", "2026-06-30", 1550, 3100, 48.0, "EXP-COM-A"],
        
        # Test Case: Missing Account No (Required)
        ["", "Asha Foods Ltd", "COM", "102", "1192834", "REF-202606-04", "2026-06-01", "2026-06-30", 400, 1500, 45.0, "EXP-COM-A"],
        
        # Test Case: Invalid Date format / Missing Date
        ["2009871032", "Royal College Colombo", "BOC", "001", "9028104", "REF-202606-05", "", "2026-06-30", 600, 1200, 38.0, "EXP-GOV-C"]
    ],
    "Kandy Central": [
        ["1008761230", "Dilmah Tea Factory", "COM", "104", "1209384", "REF-202606-06", "2026-06-01", "2026-06-30", 950, 5100, 48.0, "EXP-IND-A"],
        ["4001928304", "Kandy City Centre", "SAMP", "002", "4492018", "REF-202606-07", "2026-06-01", "2026-06-30", 4500, 6800, 52.0, "EXP-COM-B"],
        
        # Duplicate record check (same customer account, ref, and period as Colombo West's Sun Industrial record)
        # Note: If Sun Industrial was already processed, uploading this row checks if the database prevents duplicate entries
        ["1002345091", "Sun Industrial Pvt Ltd", "BOC", "032", "7045920", "REF-202606-01", "2026-06-01", "2026-06-30", 1100, 6400, 48.0, "EXP-IND-A"]
    ],
    "Galle South": [
        ["5009871230", "Jetwing Lighthouse Galle", "HNB", "008", "9928173", "REF-202606-08", "2026-06-01", "2026-06-30", 7200, 9900, 52.0, "EXP-COM-B"],
        ["2004561001", "Lanka Hospitals Corp (Galle)", "HNB", "001", "4592810", "REF-202606-09", "2026-06-01", "2026-06-30", 2200, 1900, 52.0, "EXP-MED-B"],
        
        # Test Case: Invalid bank code (warning only - should still import successfully)
        ["6008761042", "Galle Marina Resort", "XYZ", "020", "1122334", "REF-202606-10", "2026-06-01", "2026-06-30", 800, 1500, 48.0, "EXP-COM-A"],
        
        # Test Case: Negative Imports (Error - should skip)
        ["6008761043", "Galle Port Authority", "COM", "102", "8899221", "REF-202606-11", "2026-06-01", "2026-06-30", -250, 1000, 45.0, "EXP-GOV-B"],
        
        # Test Case: Negative Unit Cost (Error - should skip)
        ["6008761044", "Galle Fisheries Co", "BOC", "005", "4455667", "REF-202606-12", "2026-06-01", "2026-06-30", 300, 1200, -5.0, "EXP-IND-B"]
    ]
}

# Style options
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Segoe UI", size=10)
fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

for sheet_name, rows in sheets_data.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    # Write header
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    
    # Write data rows
    for row in rows:
        ws.append(row)
        curr_row = ws.max_row
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = font_data
            cell.alignment = align_left if col_idx in [2, 12] else align_center

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save file
output_path = "c:\\Users\\janos\\OneDrive\Desktop\\CEB demo\\ceb_billing_test_data.xlsx"
wb.save(output_path)
print(f"Spreadsheet generated successfully at: {output_path}")
